import io
from flask import Blueprint, request, render_template, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from utils import fetch_all, fetch_one, clean

consultas_bp = Blueprint('consultas', __name__)


# =========================================================
# QUERY BUILDER
# =========================================================

def _build_query(cc_id, area_id, texto, estatus, asignada, con_equipo):
    where = []
    params = []

    if cc_id:
        where.append("cc.id = %s")
        params.append(cc_id)

    if area_id:
        where.append("a.id = %s")
        params.append(area_id)

    if texto:
        where.append("""(
            CONCAT(e.nombre, ' ', COALESCE(e.apellido, '')) LIKE %s
            OR CAST(l.numero AS CHAR) LIKE %s
            OR e.correo LIKE %s
        )""")
        like = f"%{texto}%"
        params.extend([like, like, like])

    if estatus:
        where.append("l.estatus = %s")
        params.append(estatus)

    if asignada == "si":
        where.append("asi.id IS NOT NULL AND asi.estatus = 'Vigente'")
    elif asignada == "no":
        where.append("asi.id IS NULL")

    if con_equipo == "con":
        where.append("eq.imei IS NOT NULL")
    elif con_equipo == "sin":
        where.append("eq.imei IS NULL")

    where_clause = "WHERE " + " AND ".join(where) if where else ""
    return where_clause, params


def _run_consulta(cc_id, area_id, texto, estatus, asignada, con_equipo):
    where_clause, params = _build_query(cc_id, area_id, texto, estatus, asignada, con_equipo)

    sql = f"""
        SELECT
            IF(l.lada = 0, CAST(l.numero AS CHAR), CONCAT(l.lada, '-', l.numero)) AS telefono,
            l.plan,
            l.estatus,
            cc.nombre                                               AS centro_costo,
            CASE
                WHEN asi.id IS NOT NULL THEN 'Vigente'
                ELSE 'No asignada'
            END                                                     AS tipo_asignacion,
            CONCAT(e.nombre, ' ', COALESCE(e.apellido, ''))         AS empleado_nombre,
            e.correo,
            ar.nombre                                               AS area_nombre,
            eq.imei                                                 AS equipo_imei,
            eq.modelo                                               AS equipo_modelo,
            eq.serial                                               AS equipo_serial
        FROM lineas l
        LEFT JOIN cuentas_padre cp ON cp.id = l.cuenta_padre_id
        LEFT JOIN asignaciones asi ON asi.linea_id = l.id AND asi.estatus = 'Vigente' AND asi.fecha_fin IS NULL
        LEFT JOIN empleados e ON e.id = asi.empleado_id
        LEFT JOIN areas ar ON ar.id = e.area_id
        LEFT JOIN centros_costo cc ON cc.id = ar.centro_costo_id
        LEFT JOIN equipos eq ON eq.imei = asi.imei
        {where_clause}
        ORDER BY cc.nombre, l.lada, l.numero
        LIMIT 1000
    """
    return fetch_all(sql, params or None)


# =========================================================
# RUTAS
# =========================================================

@consultas_bp.route("/consultas", methods=["GET"])
def consultas():
    cc_id = clean(request.args.get("cc_id", ""))
    area_id = clean(request.args.get("area_id", ""))
    texto = clean(request.args.get("texto", ""))
    estatus = clean(request.args.get("estatus", ""))
    asignada = request.args.get("asignada", "todos")
    con_equipo = request.args.get("con_equipo", "todos")

    # Si el usuario sometió el formulario (hay cualquier parámetro), ejecutamos la consulta
    # — aunque todos los filtros estén en "todos" (devuelve todo el universo, limitado a 1000).
    # Solo mostramos el estado vacío inicial cuando se entra directo sin parámetros.
    hay_busqueda = bool(request.args)
    resultados = _run_consulta(cc_id, area_id, texto, estatus, asignada, con_equipo) if hay_busqueda else []

    centros = fetch_all("SELECT id, nombre FROM centros_costo ORDER BY nombre")

    areas = []
    if cc_id:
        areas = fetch_all(
            "SELECT id, nombre FROM areas WHERE centro_costo_id = %s ORDER BY nombre",
            (cc_id,)
        )

    return render_template(
        "consultas.html",
        resultados=resultados,
        centros=centros,
        areas=areas,
        filtros={
            "cc_id": cc_id,
            "area_id": area_id,
            "texto": texto,
            "estatus": estatus,
            "asignada": asignada,
            "con_equipo": con_equipo,
        },
        total=len(resultados),
    )


@consultas_bp.route("/api/areas_por_cc", methods=["GET"])
def api_areas_por_cc():
    cc_id = clean(request.args.get("cc_id", ""))
    if not cc_id:
        return jsonify([])
    areas = fetch_all(
        "SELECT id, nombre FROM areas WHERE centro_costo_id = %s ORDER BY nombre",
        (cc_id,)
    )
    return jsonify(areas)


@consultas_bp.route("/consultas/excel", methods=["GET"])
def descargar_excel():
    cc_id = clean(request.args.get("cc_id", ""))
    area_id = clean(request.args.get("area_id", ""))
    texto = clean(request.args.get("texto", ""))
    estatus = clean(request.args.get("estatus", ""))
    asignada = request.args.get("asignada", "todos")
    con_equipo = request.args.get("con_equipo", "todos")

    rows = _run_consulta(cc_id, area_id, texto, estatus, asignada, con_equipo)

    wb = Workbook()
    ws = wb.active
    ws.title = "Consulta de Líneas"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    headers = [
        "Teléfono", "Plan", "Estado", "Centro de Costo",
        "Tipo Asignación", "Empleado", "Correo",
        "Área", "IMEI Equipo", "Modelo Equipo", "Serial Equipo"
    ]
    ws.append(headers)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.row_dimensions[1].height = 20

    for row in rows:
        ws.append([
            row.get("telefono") or "",
            row.get("plan") or "",
            row.get("estatus") or "",
            row.get("centro_costo") or "",
            row.get("tipo_asignacion") or "",
            row.get("empleado_nombre") or "",
            row.get("correo") or "",
            row.get("area_nombre") or "",
            row.get("equipo_imei") or "",
            row.get("equipo_modelo") or "",
            row.get("equipo_serial") or "",
        ])

    col_widths = [16, 20, 15, 25, 15, 30, 30, 25, 18, 20, 20]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="consulta_lineas.xlsx",
    )
